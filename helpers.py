'''
Helper functions.
'''


import shutil
from openpyxl.utils.cell import _get_column_letter as get_column_letter

#TODO: Header to Letters

def check_max_parameter(parameter, csv_data, max_value=1):
    try:
        if max(csv_data[parameter]) >= max_value:
            return "Max " + raise_warning(
                                    parameter,
                                    report_value = max(csv_data[parameter])
                                    )
    except KeyError:
        return no_key(parameter)
    else:
        return ''

def check_min_parameter(parameter, csv_data, min_value=1):
    try:
        if min(csv_data[parameter]) <= min_value:
            return "Min " + raise_warning(
                                    parameter,
                                    report_value = min(csv_data[parameter])
                                    )
    except KeyError:
        return no_key(parameter)
    else:
        return ''

def check_yes_parameter(parameter, csv_data):
    try:
        if "Yes" in csv_data[parameter]:
            return raise_warning(parameter, report_value="Yes")
    except KeyError:
        return no_key(parameter)
    else:
        return ''

def check_ripple(parameter, csv_data):
    '''
    "The ripple limits, according to the ATX specification,
    are 120mV for the +12V and -12V rails,
    and 50mV for the remaining rails (5V, 3.3V and 5VSB).
    '''
    max_ripple_values = {
                            "+12V [V]": .12,
                            "+5V [V]": .05,
                            "+3.3V [V]": .05
    }
    count = 1

    try:
        avg = sum(csv_data[parameter]) / len(csv_data[parameter])
    except KeyError:
        return no_key(parameter)
    else:
        for i in csv_data[parameter]:
            if abs(i - avg) > max_ripple_values[parameter]:
                count += 1

        if len(csv_data[parameter]) / 100 * 80 <= count:
            return "Ripple on {}.\n".format(parameter)

def raise_warning(parameter, report_value=''):
    return "{0}: {1}\n".format(parameter, report_value)

def no_key(parameter):
    return "No data: {}\n".format(parameter)

def find_header_column_string(header_list, parameter):
    return get_column_letter(header_list.index(parameter) + 1)

def make_html(sorted_msg_dict):
    htmldoc = "<h1> Here you are, your results.</h1><br>\n"
    for key, msg in sorted_msg_dict.items():
        htmldoc += "<h2>{}</h2>\n".format(key.title())
        htmldoc += "<p>"
        for line in msg:
            htmldoc += "{}<br>".format(line.title())
        if msg == []:
            htmldoc += "Nothing to display.<br>"

        htmldoc += "</p>"

    return htmldoc

#Put all "no data" entries to the back. Put all entries starting with "--"
#into the middle.

def sort_message(msg):
    unsorted_msg = msg.split("\n")
    for idx, part in enumerate(unsorted_msg):
        if not part.startswith("No data"):
            unsorted_msg.insert(0, unsorted_msg.pop(idx))

    sorted_msg = [x for x in unsorted_msg if x != '']

    return sorted_msg
